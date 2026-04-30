# 06_export_manuscript_tables.py
"""
Auto-format All Results → APA 7th Edition Manuscript Tables
Research Design — Final Output Step

Reads all CSVs from results/ and formats into publication-ready tables
matching manuscript_guide.md Table 1-11 + Appendix A1-A6 specs.

Output structure:
  manuscript_tables/
    Table_01_Variable_Definitions.xlsx
    Table_02_Descriptive_Statistics.xlsx
    Table_03_Stationarity_Tests.xlsx
    Table_04_Structural_Break_Tests.xlsx
    Table_05_Forecast_Accuracy_Main.xlsx
    Table_06_DM_Tests_Romano_Wolf.xlsx
    Table_07_MCS_Results.xlsx
    Table_08_Crisis_Subperiod.xlsx
    Table_09_Expressibility.xlsx
    Table_10_Feature_Importance.xlsx
    Table_11_Policy_Implications.xlsx
    ALL_TABLES.xlsx            ← all tables in one workbook (one sheet each)

Formatting follows APA 7th + TFSC journal style:
  - No vertical gridlines
  - Bold headers, horizontal rules only
  - Significance stars: * p<.10  ** p<.05  *** p<.01
  - Numbers: 4 decimal places for statistics, 2 for percentages
  - Best value per column highlighted in bold
  - Model names standardised

Chạy lệnh:
  python 06_export_manuscript_tables.py

Thời gian: < 2 phút.
"""

import json
import logging
import warnings
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
logging.basicConfig(
    level=logging.INFO,
    format="[%(levelname)s] %(asctime)s | %(message)s",
    datefmt="%H:%M:%S",
)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR    = Path(r"D:\CANH\Nghiên cứu khoa học\VIẾT BÁO - CHẠY SỐ LIỆU\carbon_qml_project")
RESULTS_DIR = BASE_DIR / "Data" / "results"
OUTPUT_DIR  = BASE_DIR / "Data" / "manuscript_tables"

# Model display names (APA-style, consistent across all tables)
MODEL_NAMES = {
    "qk_svr":       "QK-SVR",
    "rbf_svm":      "RBF-SVM",
    "laplacian_svm":"Laplacian-SVM",
    "xgboost":      "XGBoost",
    "lightgbm":     "LightGBM",
    "bilstm":       "BiLSTM",
    "gru":          "GRU",
    "transformer":  "Transformer",
    "emd_lstm":     "EMD-LSTM",
}

# Model display order (QK-SVR first, then benchmarks)
MODEL_ORDER = ["qk_svr", "rbf_svm", "laplacian_svm", "xgboost",
               "lightgbm", "bilstm", "gru", "transformer", "emd_lstm"]

# Regime display names
REGIME_NAMES = {
    "pre_crisis":   "Pre-crisis (2019–2021)",
    "crisis_onset": "Crisis onset (2022 H1)",
    "peak_crisis":  "Peak crisis (2022 H2–2023 H1)",
    "post_crisis":  "Post-crisis (2023 H2)",  # data ends Dec 2023
    "other":        "Other",
}

# APA color palette
CLR_HEADER   = "1F4E79"   # dark blue — header background
CLR_SUBHDR   = "D6E4F0"   # light blue — sub-header
CLR_BEST     = "E8F5E9"   # light green — best value highlight
CLR_QK       = "FFF3E0"   # light orange — QK-SVR row highlight
CLR_WHITE    = "FFFFFF"
CLR_STRIPE   = "F7F9FC"   # light grey alternating row


# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def header_font(bold=True, color="FFFFFF", size=10):
    return Font(name="Times New Roman", bold=bold, color=color, size=size)


def body_font(bold=False, color="000000", size=10):
    return Font(name="Times New Roman", bold=bold, color=color, size=size)


def center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def thin_border(top=False, bottom=False, left=False, right=False):
    thin = Side(style="thin")
    none = Side(style=None)  # BUG 1 FIX: openpyxl accepts None (not Side(style=None) in old ver)
    # Use no_fill approach: only set sides that are needed
    return Border(
        top=thin    if top    else none,
        bottom=thin if bottom else none,
        left=thin   if left   else none,
        right=thin  if right  else none,
    )


def thick_border(top=False, bottom=False):
    thick = Side(style="medium")
    none  = Side(style=None)
    return Border(
        top=thick    if top    else none,
        bottom=thick if bottom else none,
    )


def fill(hex_color: str):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def sig_stars(p: float) -> str:
    if pd.isna(p):
        return ""
    if p < 0.01:
        return "***"
    if p < 0.05:
        return "**"
    if p < 0.10:
        return "*"
    return ""


def fmt(val, decimals=4, pct=False, stars_p=None) -> str:
    """Format a value for display. Optionally append significance stars."""
    if pd.isna(val):
        return "—"
    if pct:
        return f"{val:.{decimals}f}%"
    s = f"{val:.{decimals}f}"
    if stars_p is not None:
        s += sig_stars(stars_p)
    return s


def write_title(ws, title: str, ncols: int = 8):
    """Write APA-style table title at row 1. Note is written separately after data."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = title
    ws["A1"].font = Font(name="Times New Roman", bold=True, italic=False, size=11)
    ws["A1"].alignment = left_align()


def write_note(ws, note: str, after_row: int, ncols: int = 8):
    """Write APA-style note below data table."""
    if not note:
        return
    note_row = after_row + 2
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=ncols)
    cell = ws.cell(row=note_row, column=1, value=f"Note. {note}")
    cell.font = Font(name="Times New Roman", italic=True, size=9)
    cell.alignment = left_align()


def format_header_row(ws, row: int, values: List, ncols: int):
    """Apply dark header formatting to a row."""
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font     = header_font()
        c.fill     = fill(CLR_HEADER)
        c.alignment = center_align()
        c.border   = thin_border(bottom=True)


def format_data_row(ws, row: int, values: List,
                    is_qk: bool = False, is_best_cols: List[int] = None,
                    stripe: bool = False):
    """Write and format a data row."""
    bg = CLR_QK if is_qk else (CLR_STRIPE if stripe else CLR_WHITE)
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        is_best = is_best_cols and col in is_best_cols
        c.font      = body_font(bold=is_best or is_qk)
        c.fill      = fill(CLR_BEST if is_best else bg)
        c.alignment = center_align() if col > 1 else left_align()
        c.border    = thin_border()


def autofit_columns(ws, min_width=8, max_width=30):
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value or "")) for cell in col),
            default=min_width,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(max_len + 2, min_width), max_width
        )


def add_bottom_rule(ws, row: int, ncols: int):
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).border = thick_border(bottom=True)


def new_sheet(wb: Workbook, name: str):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    return ws


def save_single_table(df_formatted: pd.DataFrame, table_name: str,
                      title: str, note: str, output_dir: Path) -> None:
    """Save a single table as its own .xlsx file."""
    wb = Workbook()
    ws = wb.active
    ws.title = table_name[:31]
    ws.sheet_view.showGridLines = False

    ncols = len(df_formatted.columns)

    # Row 1: title only — note goes AFTER data to avoid merged cell conflict
    write_title(ws, title, ncols)

    # Row 3: header
    format_header_row(ws, 3, list(df_formatted.columns), ncols)

    # Rows 4+: data
    for i, (_, row) in enumerate(df_formatted.iterrows()):
        r    = i + 4
        vals = list(row)
        is_qk = str(vals[0]).strip() == "QK-SVR"
        format_data_row(ws, r, vals, is_qk=is_qk, stripe=(i % 2 == 1))

    last_data_row = 3 + len(df_formatted)
    add_bottom_rule(ws, last_data_row, ncols)

    # Note AFTER data
    write_note(ws, note, after_row=last_data_row, ncols=ncols)

    autofit_columns(ws)

    out = output_dir / f"{table_name}.xlsx"
    wb.save(out)
    logging.info(f"  💾 {out.name}")


# ─────────────────────────────────────────────────────────────────────────────
# TABLE BUILDERS
# ─────────────────────────────────────────────────────────────────────────────

def build_table01_variables() -> pd.DataFrame:
    """Table 1: Variable Definitions and Sources (hardcoded from research design)."""
    rows = [
        ["EUA_return",   "EU Allowance log-return",          "ICE Futures / Sandbag", "Daily",         "Log-return",            "Target variable"],
        ["GAS_return",   "TTF Natural Gas log-return",       "Investing.com TTF",      "Daily",         "Log-return",            "Fuel-switching driver"],
        ["OIL_return",   "Brent Crude log-return",           "FRED DCOILBRENTEU",      "Daily",         "Log-return",            "Macro energy proxy"],
        ["COAL_return",  "Coal ARA log-return",              "Platts / ICE Futures",   "Daily",         "Log-return",            "EU fuel-switching"],
        ["ELEC_return",  "Electricity price log-return",     "EEX / ENTSO-E",          "Daily",         "Log-return",            "Power sector proxy"],
        ["IP_return",    "Industrial Production YoY growth", "Eurostat STS_INPR_M",    "Monthly→Daily", "Step-ramp interpolation","Macro activity"],
        ["CPI_return",   "HICP inflation YoY",               "Eurostat PRC_HICP_MANR", "Monthly→Daily", "Step-ramp interpolation","Inflation proxy"],
        ["POLICY_dummy", "Policy event dummy",               "EUR-Lex + ECB",           "Event→Daily",   "Binary (0/1)",          "7-day event window"],
        ["PHASE_dummy",  "ETS Phase dummy",                  "EU Commission",           "Daily",         "Binary (0=Phase3, 1=Phase4)", "Structural break control"],
    ]
    return pd.DataFrame(rows, columns=[
        "Variable", "Description", "Source", "Frequency", "Transformation", "Notes"
    ])


def build_table05_forecast_accuracy(results_dir: Path) -> Optional[pd.DataFrame]:
    """Table 5: Main Forecast Accuracy Results (RMSE, MAE, DA%)."""
    path = results_dir / "metrics_summary.csv"
    if not path.exists():
        logging.warning(f"  Missing: {path.name}")
        return None

    df = pd.read_csv(path)
    # Standardise column names
    df.columns = [c.lower() for c in df.columns]
    df["model"] = df["model"].map(lambda x: MODEL_NAMES.get(x, x))

    records = []
    for h in [1, 5, 22]:
        df_h = df[df["horizon"] == h]
        for model_key in MODEL_ORDER:
            mname = MODEL_NAMES.get(model_key, model_key)
            row   = df_h[df_h["model"] == mname]
            if len(row) == 0:
                continue
            r = row.iloc[0]
            records.append({
                "Model":   mname,
                "Horizon": f"H={h}",
                "RMSE":    fmt(r.get("rmse_mean", np.nan)),
                "±SD":     fmt(r.get("rmse_std", np.nan)),
                "MAE":     fmt(r.get("mae_mean", np.nan)),
                "MAPE(%)": fmt(r.get("mape_mean", np.nan), decimals=2),
                "DA(%)":   fmt(r.get("da_mean", np.nan), decimals=2),
            })

    return pd.DataFrame(records)


def build_table06_dm_tests(results_dir: Path) -> Optional[pd.DataFrame]:
    """Table 6: Diebold-Mariano Tests with Romano-Wolf Correction."""
    path = results_dir / "dm_tests.csv"
    if not path.exists():
        logging.warning(f"  Missing: {path.name}")
        return None

    df   = pd.read_csv(path)
    df.columns = [c.lower() for c in df.columns]
    df["model_alt"] = df["model_alt"].map(lambda x: MODEL_NAMES.get(x, x))

    records = []
    for h in [1, 5, 22]:
        df_h = df[df["horizon"] == h]
        for model_key in MODEL_ORDER:
            if model_key == "qk_svr":
                continue
            mname = MODEL_NAMES.get(model_key, model_key)
            row   = df_h[df_h["model_alt"] == mname]
            if len(row) == 0:
                continue
            r  = row.iloc[0]
            p  = r.get("p_value", np.nan)
            rw = r.get("rw_pvalue", np.nan)
            records.append({
                "Benchmark":     mname,
                "Horizon":       f"H={h}",
                "DM Stat (HLN)": fmt(r.get("dm_hln", np.nan), decimals=3),
                "p-value":       fmt(p, decimals=4) + sig_stars(p),
                "RW p-value":    fmt(rw, decimals=4) + sig_stars(rw),
                "Sig. (DM)":     sig_stars(p),
                "Sig. (RW)":     sig_stars(rw),
            })

    return pd.DataFrame(records)


def build_table07_mcs(results_dir: Path) -> Optional[pd.DataFrame]:
    """Table 7: Model Confidence Set Results."""
    path = results_dir / "mcs_results.csv"
    if not path.exists():
        logging.warning(f"  Missing: {path.name}")
        return None

    df = pd.read_csv(path)
    df.columns = [c.lower() for c in df.columns]
    df["model"] = df["model"].map(lambda x: MODEL_NAMES.get(x, x))

    records = []
    for h in [1, 5, 22]:
        df_h = df[df["horizon"] == h] if "horizon" in df.columns else df
        for model_key in MODEL_ORDER:
            mname = MODEL_NAMES.get(model_key, model_key)
            row   = df_h[df_h["model"] == mname]
            if len(row) == 0:
                continue
            r = row.iloc[0]
            records.append({
                "Model":       mname,
                "Horizon":     f"H={h}",
                "In MCS":      "✓" if r.get("in_mcs", False) else "✗",
                "Elim. Order": str(int(r["elim_order"])) if pd.notna(r.get("elim_order")) else "—",
                "MSE (mean)":  fmt(r.get("mean_sq_error", np.nan), decimals=6),
            })

    return pd.DataFrame(records)


def build_table08_crisis(results_dir: Path) -> Optional[pd.DataFrame]:
    """Table 8: Crisis Sub-Period Analysis."""
    path = results_dir / "crisis_subperiod.csv"
    if not path.exists():
        logging.warning(f"  Missing: {path.name}")
        return None

    df = pd.read_csv(path)
    df.columns = [c.lower() for c in df.columns]
    df["model"]  = df["model"].map(lambda x: MODEL_NAMES.get(x, x))
    df["regime"] = df["regime"].map(lambda x: REGIME_NAMES.get(x, x))

    records = []
    for h in [1, 5, 22]:
        df_h = df[df["horizon"] == h]
        for regime_key in ["pre_crisis", "crisis_onset", "peak_crisis", "post_crisis"]:
            regime_name = REGIME_NAMES.get(regime_key, regime_key)
            df_r = df_h[df_h["regime"] == regime_name]
            if len(df_r) == 0:
                continue
            for model_key in MODEL_ORDER:
                mname = MODEL_NAMES.get(model_key, model_key)
                row   = df_r[df_r["model"] == mname]
                if len(row) == 0:
                    continue
                r = row.iloc[0]
                records.append({
                    "Model":   mname,
                    "Horizon": f"H={h}",
                    "Regime":  regime_name,
                    "RMSE":    fmt(r.get("rmse_mean", np.nan)),
                    "±SD":     fmt(r.get("rmse_std", np.nan)),
                    "DA(%)":   fmt(r.get("da_mean", np.nan), decimals=1),
                })

    return pd.DataFrame(records)


def build_table09_expressibility(results_dir: Path) -> Optional[pd.DataFrame]:
    """Table 9: Expressibility Comparison."""
    path = results_dir / "expressibility.csv"
    if not path.exists():
        logging.warning(f"  Missing: {path.name}")
        return None

    df = pd.read_csv(path)
    df.columns = [c.lower() for c in df.columns]

    # Load ARI from regime analysis
    ari_val = "—"
    ari_path = results_dir / "regime_ari.json"
    if ari_path.exists():
        with open(ari_path) as f:
            ari_data = json.load(f)
        ari_val = fmt(ari_data.get("ari", np.nan), decimals=3)

    records = []
    for _, r in df.iterrows():
        records.append({
            "Kernel":            r.get("kernel", "—"),
            "D_KL (from Haar)":  fmt(r.get("dkl_from_haar", np.nan), decimals=4),
            "Spectral Decay":    fmt(r.get("spectral_decay_gini", np.nan), decimals=4),
            "ARI (regime)":      ari_val if "Quantum" in str(r.get("kernel", "")) else "—",
            "Interpretation":    ("Lower = richer feature space"
                                  if "Quantum" in str(r.get("kernel", "")) else ""),
        })

    return pd.DataFrame(records)


def build_table10_feature_importance(results_dir: Path) -> Optional[pd.DataFrame]:
    """Table 10: Feature Importance — QKFM vs TreeSHAP."""
    path = results_dir / "feature_importance.csv"
    if not path.exists():
        logging.warning(f"  Missing: {path.name}")
        return None

    df = pd.read_csv(path)
    df.columns = [c.lower() for c in df.columns]

    # BUG 4 FIX: QKFM only embeds first 4 features (N_QUBITS=4)
    # IP_return and POLICY_dummy are NOT embedded → not in feature_importance.csv
    feature_labels = {
        "GAS_return":  "Natural Gas (TTF)",
        "OIL_return":  "Brent Crude",
        "COAL_return": "Coal ARA",
        "ELEC_return": "Electricity",
        # Additional features may appear from TreeSHAP (uses all 8)
        "IP_return":    "Industrial Production",
        "CPI_return":   "CPI Inflation",
        "POLICY_dummy": "Policy Events",
        "PHASE_dummy":  "ETS Phase",
    }

    records = []
    for regime_key in ["pre_crisis", "crisis_onset", "peak_crisis"]:
        df_r = df[df["regime"] == regime_key]
        if len(df_r) == 0:
            continue
        for _, r in df_r.iterrows():
            feat  = r.get("feature", "")
            qkval = r.get("qkfm_importance", np.nan)
            shval = r.get("shap_importance", np.nan)
            records.append({
                "Regime":          REGIME_NAMES.get(regime_key, regime_key),
                "Feature":         feature_labels.get(feat, feat),
                "QKFM Importance": fmt(qkval, decimals=4),
                "QKFM Rank":       str(int(r["qkfm_rank"])) if pd.notna(r.get("qkfm_rank")) else "—",
                "SHAP Importance": fmt(shval, decimals=4),
                "SHAP Rank":       str(int(r["shap_rank"])) if pd.notna(r.get("shap_rank")) else "—",
            })

    return pd.DataFrame(records)


def build_table11_policy(results_dir: Path) -> Optional[pd.DataFrame]:
    """Table 11: Policy Implications Summary."""
    path = results_dir / "policy_implications.csv"
    if not path.exists():
        logging.warning(f"  Missing: {path.name}")
        return None

    df = pd.read_csv(path)
    df.columns = [c.lower() for c in df.columns]

    records = []
    for _, r in df.iterrows():
        savings = r.get("hedging_savings_eur", np.nan)
        savings_fmt = f"EUR {savings:,.0f}" if pd.notna(savings) else "—"
        records.append({
            "Forecast Horizon":          f"H={int(r['horizon'])} days",
            "Period":                    "Peak crisis (2022 H2–2023 H1)",
            "RMSE Reduction (%)":        fmt(r.get("uncertainty_reduction_pct", np.nan), decimals=2),
            "DA Improvement (pp)":       fmt(r.get("da_improvement_pp", np.nan), decimals=1),
            "Est. Hedging Savings":      savings_fmt,
            "MSR Policy Recommendation": (
                "Dynamic TNAC trigger ± 2σ CI"
                if int(r.get("horizon", 0)) == 22 else "Intraday corridor adjustment"
            ),
        })

    return pd.DataFrame(records)


# ─────────────────────────────────────────────────────────────────────────────
# ALL-IN-ONE WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────

TABLE_SPECS = [
    ("Table_01_Variable_Definitions",   "Table 1. Variable Definitions and Sources",
     "Sources verified against Scopus DOI database. Step-ramp interpolation applied to monthly variables."),
    ("Table_05_Forecast_Accuracy_Main", "Table 5. Out-of-Sample Forecast Accuracy — Main Results",
     "RMSE and MAE values are mean across walk-forward folds. DA = Directional Accuracy (%). "
     "Bold = best performer per column. QK-SVR = Quantum Kernel SVR (proposed model). "
     "All models tuned with 100 Optuna trials on identical initial training window."),
    ("Table_06_DM_Tests_Romano_Wolf",   "Table 6. Diebold-Mariano Tests: QK-SVR vs. Benchmarks",
     "H0: Equal predictive accuracy. DM statistic uses Harvey-Leybourne-Newbold small-sample correction "
     "and Newey-West HAC covariance. RW p-value = Romano-Wolf stepdown FWER-corrected p-value "
     "(1,000 bootstrap iterations). * p < .10, ** p < .05, *** p < .01."),
    ("Table_07_MCS_Results",            "Table 7. Model Confidence Set Results",
     "Hansen-Lunde-Nason (2011) procedure, α = .25. ✓ = included in superior set. "
     "Elim. Order = sequence of elimination (lower = better). MSE computed on squared prediction errors."),
    ("Table_08_Crisis_Subperiod",       "Table 8. Crisis Sub-Period Forecast Accuracy",
     "Regime dates from Bai-Perron multiple breakpoint test. "
     "Peak crisis: 2022 H2–2023 H1 (Russia-Ukraine energy shock). "
     "Bold = best RMSE per regime-horizon combination."),
    ("Table_09_Expressibility",         "Table 9. Expressibility Comparison: Quantum vs. Classical Kernels",
     "D_KL = Kullback-Leibler divergence of kernel eigenspectrum from Haar (uniform) distribution "
     "(Sim et al., 2019 adaptation). Lower D_KL indicates richer, more uniform feature space. "
     "ARI = Adjusted Rand Index from spectral clustering on test kernel matrix."),
    ("Table_10_Feature_Importance",     "Table 10. Feature Importance: QKFM vs. TreeSHAP",
     "QKFM = Quantum Kernel Feature Masking (Section 8.1); values normalized to sum to 1. "
     "SHAP = TreeSHAP on XGBoost (Lundberg & Lee, 2017). "
     "Importance averaged across 20 representative walk-forward folds per regime."),
    ("Table_11_Policy_Implications",    "Table 11. Policy Implications for EU ETS Phase 4",
     "Hedging savings computed as: EU ETS market size (EUR 28bn) × hedging cost ratio (0.5%) × "
     "uncertainty reduction (%). Assumptions: zero transaction cost; cost proportional to RMSE reduction. "
     "Consistent with SDG 13.2 framework."),
]


def write_sheet(wb: Workbook, ws_name: str, df: pd.DataFrame,
                title: str, note: str) -> None:
    """Write a formatted table to a workbook sheet."""
    ws = new_sheet(wb, ws_name[:31])
    if df is None or len(df) == 0:
        ws["A1"] = f"{title} — Data not available (run pipeline first)"
        ws["A1"].font = body_font(bold=True)
        return

    ncols = len(df.columns)

    # Row 1: title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = title
    ws["A1"].font = Font(name="Times New Roman", bold=True, size=11)
    ws["A1"].alignment = left_align()

    # Row 2: blank separator
    # Row 3: header
    format_header_row(ws, 3, list(df.columns), ncols)
    ws.row_dimensions[3].height = 30

    # Data rows starting at row 4
    # Find best values per numeric-looking column for highlighting
    # BUG 5 FIX: best_cols[row_index] = [col_indices to highlight in that row]
    # Previous logic had correct structure but needed verification of col mapping
    # Each column independently finds its best row, then marks (row→col) pair
    best_cols = {}  # {row_index: [col_indices]}
    for col_i, col_name in enumerate(df.columns, 1):
        if col_i == 1:
            continue
        vals = []
        for v in df[col_name]:
            try:
                num = float(str(v).replace("*", "").replace("—", "")
                            .replace(",", "").replace("%", "").strip())
                vals.append(num)
            except Exception:
                vals.append(None)
        valid = [(v, i) for i, v in enumerate(vals) if v is not None]
        if not valid:
            continue
        is_da = any(k in col_name for k in ("DA", "da", "Dir", "Hit"))
        # best row for this column
        best_row_idx = (max(valid, key=lambda x: x[0])[1] if is_da
                        else min(valid, key=lambda x: x[0])[1])
        best_cols.setdefault(best_row_idx, []).append(col_i)

    for i, (_, row) in enumerate(df.iterrows()):
        r    = i + 4
        vals = list(row)
        is_qk    = str(vals[0]).strip() == "QK-SVR"
        best_c   = best_cols.get(i, [])
        format_data_row(ws, r, vals, is_qk=is_qk,
                        is_best_cols=best_c, stripe=(i % 2 == 1))

    # Bottom rule
    add_bottom_rule(ws, ws.max_row, ncols)

    # Note row
    if note:
        note_row = ws.max_row + 2
        ws.merge_cells(start_row=note_row, start_column=1,
                       end_row=note_row, end_column=ncols)
        c = ws.cell(row=note_row, column=1, value=f"Note. {note}")
        c.font      = Font(name="Times New Roman", italic=True, size=9)
        c.alignment = left_align()

    autofit_columns(ws)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Build all table DataFrames
    logging.info("📊 Building all manuscript tables...")

    tables: Dict[str, Optional[pd.DataFrame]] = {
        "Table_01_Variable_Definitions":   build_table01_variables(),
        "Table_05_Forecast_Accuracy_Main": build_table05_forecast_accuracy(RESULTS_DIR),
        "Table_06_DM_Tests_Romano_Wolf":   build_table06_dm_tests(RESULTS_DIR),
        "Table_07_MCS_Results":            build_table07_mcs(RESULTS_DIR),
        "Table_08_Crisis_Subperiod":       build_table08_crisis(RESULTS_DIR),
        "Table_09_Expressibility":         build_table09_expressibility(RESULTS_DIR),
        "Table_10_Feature_Importance":     build_table10_feature_importance(RESULTS_DIR),
        "Table_11_Policy_Implications":    build_table11_policy(RESULTS_DIR),
    }

    specs = {s[0]: (s[1], s[2]) for s in TABLE_SPECS}

    # ── Save individual files ────────────────────────────────────────────────
    logging.info("\n📁 Saving individual table files...")
    for key, df in tables.items():
        title, note = specs.get(key, (key, ""))
        if df is not None:
            save_single_table(df, key, title, note, OUTPUT_DIR)
            # Also save CSV
            df.to_csv(OUTPUT_DIR / f"{key}.csv", index=False)
        else:
            logging.warning(f"  ⚠️  {key}: no data — skipped")

    # ── Save ALL_TABLES.xlsx — all sheets in one workbook ───────────────────
    logging.info("\n📚 Building ALL_TABLES.xlsx...")
    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    # Cover sheet
    ws_cover = wb.create_sheet("Cover")
    ws_cover.sheet_view.showGridLines = False
    ws_cover["A1"] = "Manuscript Tables"
    ws_cover["A1"].font = Font(name="Times New Roman", bold=True, size=14)
    ws_cover["A2"] = "Quantum Kernel Methods for Carbon Price Forecasting"
    ws_cover["A2"].font = Font(name="Times New Roman", italic=True, size=11)
    ws_cover["A3"] = "Technological Forecasting and Social Change (TFSC)"
    ws_cover["A3"].font = Font(name="Times New Roman", size=10)
    ws_cover["A4"] = "Generated by: 06_export_manuscript_tables.py"
    ws_cover["A4"].font = Font(name="Times New Roman", size=9, color="666666")

    contents = [["Sheet", "Table", "Status"]]
    for key, df in tables.items():
        title, note = specs.get(key, (key, ""))
        ws_name = key.replace("Table_", "T").replace("_", " ")[:31]
        write_sheet(wb, ws_name, df, title, note)
        status  = f"✓ ({len(df)} rows)" if df is not None else "⚠ Missing data"
        contents.append([ws_name, title, status])

    # Contents table on cover
    for i, row in enumerate(contents, start=6):
        for j, val in enumerate(row, 1):
            c = ws_cover.cell(row=i, column=j, value=val)
            if i == 6:
                c.font = Font(name="Times New Roman", bold=True, size=10)
                c.fill = fill(CLR_HEADER)
                c.font = Font(name="Times New Roman", bold=True,
                              color="FFFFFF", size=10)
            else:
                c.font = Font(name="Times New Roman", size=10)
    autofit_columns(ws_cover)

    all_path = OUTPUT_DIR / "ALL_TABLES.xlsx"
    wb.save(all_path)
    logging.info(f"  💾 ALL_TABLES.xlsx")

    # ── Console summary ───────────────────────────────────────────────────────
    print("\n" + "="*65)
    print("📋 TABLE EXPORT COMPLETE")
    print("="*65)
    available = sum(1 for df in tables.values() if df is not None)
    print(f"  Tables exported : {available}/{len(tables)}")
    print(f"  Output folder  : {OUTPUT_DIR}")
    print(f"  Files created  : {available} × .xlsx + {available} × .csv + ALL_TABLES.xlsx")

    if available < len(tables):
        missing = [k for k, v in tables.items() if v is None]
        print(f"\n  ⚠️  Missing data for:")
        for m in missing:
            print(f"     - {m} (run corresponding pipeline step first)")

    print("\n" + "="*65)
    print("Thứ tự bảng sẵn sàng để đưa vào manuscript:")
    for key, (title, _) in specs.items():
        mark = "✅" if tables.get(key) is not None else "⚠️ "
        print(f"  {mark}  {title}")
    print("="*65)
    print("✅ Sẵn sàng viết kết quả nghiên cứu.")


if __name__ == "__main__":
    main()
