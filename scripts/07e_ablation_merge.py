# =============================================================================
# 07e_ablation_merge_fixed.py  —  v2
# =============================================================================
# BUG 1 — No parquet fallback → crash nếu 07a-d saved as CSV
#   Fix: load_ablation() thử parquet trước, fallback CSV
#
# BUG 2 [CRITICAL] — Side(style=None) không hợp lệ trong openpyxl
#   → AttributeError khi render border → crash tạo Excel
#   Fix: dùng Side(style="thin") hoặc bỏ border, không dùng style=None
#
# BUG 3 — TABLE_SPECS notes nói n_seeds=30 nhưng fixed files dùng N_SEEDS=10
#   Fix: đọc n_seeds thực tế từ data thay vì hardcode trong note
#
# BUG 4 — build_summary std=NaN nếu fast mode (1 fold/config)
#   Fix: fillna(0) cho std khi n=1; hiển thị "N/A" thay vì "NaN"
# =============================================================================

import logging
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO,
                    format="[%(levelname)s] %(asctime)s | %(message)s",
                    datefmt="%H:%M:%S")

# =============================================================================
# CẤU HÌNH — chỉnh BASE_DIR nếu cần
# =============================================================================
BASE_DIR  = Path(r"D:\CANH\Nghiên cứu khoa học\VIẾT BÁO - CHẠY SỐ LIỆU\carbon_qml_project")
# =============================================================================

OUTPUT_DIR = BASE_DIR / "Data" / "results"
TABLE_DIR  = BASE_DIR / "Data" / "manuscript_tables"
TABLE_DIR.mkdir(parents=True, exist_ok=True)

BASELINES = {
    "AB1_qubits":       "n_qubits=4",
    "AB2_depth":        "n_layers=2",
    "AB3_entanglement": "entanglement=circular",
    "AB4_reuploading":  "re_upload=True",
}

ABLATION_FILES = {
    "AB1_qubits":       "ablation_ab1_qubits",
    "AB2_depth":        "ablation_ab2_depth",
    "AB3_entanglement": "ablation_ab3_entanglement",
    "AB4_reuploading":  "ablation_ab4_reuploading",
}

TABLE_SPECS = {
    "AB1_qubits": {
        "table_id": "A2",
        "title":    "Table A2. Ablation Study AB1: Effect of Qubit Count",
        "note":     ("n_layers=2, circular entanglement (fixed). "
                     "Bold = baseline configuration (n_qubits=4). "
                     "RMSE = mean ± SD across walk-forward folds."),
    },
    "AB2_depth": {
        "table_id": "A3",
        "title":    "Table A3. Ablation Study AB2: Effect of Circuit Depth",
        "note":     ("n_qubits=4, circular entanglement (fixed). "
                     "Bold = baseline (n_layers=2)."),
    },
    "AB3_entanglement": {
        "table_id": "A4",
        "title":    "Table A4. Ablation Study AB3: Effect of Entanglement Pattern",
        "note":     ("n_qubits=4, n_layers=2 (fixed). Bold = baseline (circular). "
                     "Linear: CNOT(i→i+1). Circular: CNOT(i→(i+1)%n). Full: all pairs."),
    },
    "AB4_reuploading": {
        "table_id": "A5",
        "title":    "Table A5. Ablation Study AB4: Effect of Data Re-uploading",
        "note":     ("n_qubits=4, n_layers=2, circular (fixed). "
                     "Bold = baseline (re_upload=True). "
                     "Re-uploading: AngleEmbedding at every layer "
                     "(Pérez-Salinas et al., 2020)."),
    },
}


# ── Helpers ───────────────────────────────────────────────────────
def load_ablation(stem: str) -> Optional[pd.DataFrame]:
    """BUG 1 FIX: try parquet first, fallback CSV."""
    pq  = OUTPUT_DIR / f"{stem}.parquet"
    csv = OUTPUT_DIR / f"{stem}.csv"
    if pq.exists():
        return pd.read_parquet(pq)
    if csv.exists():
        logging.info(f"  Parquet not found → loading CSV: {csv.name}")
        return pd.read_csv(csv)
    return None


def build_summary(df: pd.DataFrame, ablation_key: str) -> pd.DataFrame:
    """
    Build RMSE mean ± SD per config × horizon.
    BUG 3 FIX: n_seeds lấy từ data thực tế.
    BUG 4 FIX: std=NaN khi n=1 → hiển thị "N/A".
    """
    baseline = BASELINES.get(ablation_key, "")

    agg = (
        df.groupby(["config", "horizon"])["fold_rmse"]
        .agg(mean="mean", std="std", n="count")
        .reset_index()
    )

    # BUG 3 FIX: lấy n_seeds từ data thực tế
    if "n_seeds" in df.columns:
        seeds_per_config = df.groupby("config")["n_seeds"].median().to_dict()
    else:
        seeds_per_config = {}

    def fmt_rmse(row):
        if pd.isna(row["std"]) or row["n"] <= 1:
            return f"{row['mean']:.4f} ± N/A"  # BUG 4 FIX
        return f"{row['mean']:.4f} ± {row['std']:.4f}"

    agg["RMSE (mean ± SD)"] = agg.apply(fmt_rmse, axis=1)
    agg["Is Baseline"]      = agg["config"] == baseline
    agg["Horizon"]          = agg["horizon"].apply(lambda h: f"H={h}")
    agg["N Folds"]          = agg["n"].astype(int)
    agg["N Seeds"]          = agg["config"].map(
        lambda c: int(seeds_per_config.get(c, 0)) or "—"
    )
    agg = agg.sort_values(["config", "horizon"]).reset_index(drop=True)

    return agg[["config", "Horizon", "RMSE (mean ± SD)", "N Folds", "N Seeds", "Is Baseline"]]


# ── Excel style helpers ───────────────────────────────────────────
def hdr_font(): return Font(name="Times New Roman", bold=True, color="FFFFFF", size=10)
def body_font(bold=False): return Font(name="Times New Roman", bold=bold, size=10)
def note_font(): return Font(name="Times New Roman", italic=True, size=9)
def center(): return Alignment(horizontal="center", vertical="center")
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def fill_c(hex_c): return PatternFill("solid", start_color=hex_c, fgColor=hex_c)

def top_border():    return Border(top=Side(style="medium"))
def bottom_border(): return Border(bottom=Side(style="medium"))
def thin_border():   return Border(bottom=Side(style="thin"))

def autofit(ws, min_w=10, max_w=40):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w+2, min_w), max_w)


# ── Write xlsx ────────────────────────────────────────────────────
def write_ablation_table(df: pd.DataFrame, ablation_key: str, out_path: Path) -> None:
    spec     = TABLE_SPECS[ablation_key]
    baseline = BASELINES.get(ablation_key, "")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Table {spec['table_id']}"[:31]
    ws.sheet_view.showGridLines = False

    display_cols = ["config", "Horizon", "RMSE (mean ± SD)", "N Folds", "N Seeds"]
    col_labels   = ["Configuration", "Horizon", "RMSE (mean ± SD)", "N Folds", "N Seeds"]
    ncols        = len(display_cols)

    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"]           = spec["title"]
    ws["A1"].font      = Font(name="Times New Roman", bold=True, size=11)
    ws["A1"].alignment = left()

    # Row 3: Header
    for col, label in enumerate(col_labels, 1):
        c            = ws.cell(row=3, column=col, value=label)
        c.font       = hdr_font()
        c.fill       = fill_c("1F4E79")
        c.alignment  = center()
        # BUG 2 FIX: use valid border styles only
        c.border     = Border(top=Side(style="medium"), bottom=Side(style="thin"))
    ws.row_dimensions[3].height = 28

    # Data rows
    for i, (_, row) in enumerate(df.iterrows()):
        r       = i + 4
        is_base = bool(row.get("Is Baseline", False))
        stripe  = (i % 2 == 1)
        bg      = "FFF3E0" if is_base else ("F7F9FC" if stripe else "FFFFFF")
        for col, key in enumerate(display_cols, 1):
            c           = ws.cell(row=r, column=col, value=str(row[key]))
            c.font      = body_font(bold=is_base)
            c.fill      = fill_c(bg)
            c.alignment = left() if col == 1 else center()
            # BUG 2 FIX: only thin bottom, no None sides
            c.border    = thin_border()

    # Bottom rule after last data row
    last_row = 3 + len(df)
    for col in range(1, ncols+1):
        ws.cell(row=last_row, column=col).border = bottom_border()

    # Note row
    note_row = last_row + 2
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=ncols)
    c           = ws.cell(row=note_row, column=1, value=f"Note. {spec['note']}")
    c.font      = note_font()
    c.alignment = left()
    ws.row_dimensions[note_row].height = 50

    autofit(ws)
    wb.save(out_path)
    logging.info(f"  Saved: {out_path.name}")


# ── Main ──────────────────────────────────────────────────────────
def main():
    all_summaries = []
    available     = 0
    missing       = []

    for ablation_key, stem in ABLATION_FILES.items():
        df_raw = load_ablation(stem)
        if df_raw is None:
            logging.warning(f"⚠️  {stem} not found — skipping.")
            missing.append(ablation_key)
            continue

        logging.info(f"Processing {ablation_key}: {len(df_raw)} rows | "
                     f"configs: {df_raw['config'].unique().tolist()}")

        df_summary = build_summary(df_raw, ablation_key)

        # Write XLSX table
        spec_id  = TABLE_SPECS[ablation_key]["table_id"]
        out_name = f"Table_{spec_id}_Ablation_{ablation_key}.xlsx"
        write_ablation_table(df_summary, ablation_key, TABLE_DIR / out_name)

        # Write CSV summary
        df_summary.to_csv(OUTPUT_DIR / f"ablation_{ablation_key}_summary_v2.csv", index=False)

        df_summary["ablation"] = ablation_key
        all_summaries.append(df_summary)
        available += 1

    # Combined CSV
    if all_summaries:
        df_all = pd.concat(all_summaries, ignore_index=True)
        df_all.to_csv(OUTPUT_DIR / "ablation_all_summary.csv", index=False)
        logging.info("Saved: ablation_all_summary.csv")

    # Console report
    print(f"\n{'='*65}")
    print("  ABLATION MERGE COMPLETE")
    print(f"{'='*65}")
    print(f"  Files merged: {available}/{len(ABLATION_FILES)}")
    print(f"  Tables dir  : {TABLE_DIR}")

    if missing:
        print(f"\n  ⚠️  Chưa có kết quả từ: {missing}")
        print("     → Chạy các file 07a-d tương ứng trước.")

    if all_summaries:
        print("\n  📊 Kết quả ablation (RMSE mean):")
        df_show = pd.concat(all_summaries)[
            ["ablation", "config", "Horizon", "RMSE (mean ± SD)", "N Folds"]
        ].rename(columns={"config": "Config"})
        print(df_show.to_string(index=False))

    print(f"\n{'='*65}")
    if available == len(ABLATION_FILES):
        print("  ✅ Appendix B tables (A2-A5) sẵn sàng cho bản thảo.")
    else:
        print("  ⚠️  Một số ablation chưa chạy xong — bảng sẽ thiếu dữ liệu.")
    print(f"{'='*65}")


if __name__ == "__main__":
    main()
