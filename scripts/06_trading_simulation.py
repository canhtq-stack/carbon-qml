# =============================================================================
# 06_trading_simulation_fixed.py  —  v2
# =============================================================================
# BUG 1 [CRITICAL] — load_predictions đọc QK-SVR 2 lần → duplicate
#   Fix: chỉ đọc benchmark_predictions.parquet (đã có QK-SVR từ 03d)
#
# BUG 2 [CRITICAL] — run_simulation concatenates y_true/y_pred như list
#   nhưng sau 03d_merge_fixed chúng là scalar
#   Fix: dùng to_scalar() để convert an toàn
#
# BUG 3 [CRITICAL] — build_cumulative_returns cùng vấn đề scalar/list
#   Fix: same as BUG 2
#
# BUG 4 — Note cell có prefix "Note. " 2 lần
#   Fix: xóa prefix thừa trong f-string
#
# BUG 5 — Regime mask dùng Python loop O(N) → vectorize với numpy
#   Fix: pd.Series.between() vectorized
#
# BUG 6 — No parquet fallback
#   Fix: try parquet → fallback CSV
#
# PERF — y_pred_mean/y_pred column name: handle cả hai
# =============================================================================

import json
import logging
import warnings
from pathlib import Path
from typing import Dict, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
logging.basicConfig(
    level=logging.INFO,
    format="[%(levelname)s] %(asctime)s | %(message)s",
    datefmt="%H:%M:%S",
)

# =============================================================================
# CẤU HÌNH — chỉnh tại đây rồi nhấn Run trong VS Code
# =============================================================================
BASE_DIR = Path(r"D:\CANH\Nghiên cứu khoa học\VIẾT BÁO - CHẠY SỐ LIỆU\carbon_qml_project")
# =============================================================================

RESULTS_DIR = BASE_DIR / "Data" / "results"
CONFIG_DIR  = BASE_DIR / "config"
TABLE_DIR   = BASE_DIR / "Data" / "manuscript_tables"
TABLE_DIR.mkdir(parents=True, exist_ok=True)

TRADING_DAYS_PER_YEAR = 252
RANDOM_SEED           = 42

MODEL_NAMES = {
    "qk_svr":        "QK-SVR",
    "rbf_svm":       "RBF-SVM",
    "laplacian_svm": "Laplacian-SVM",
    "xgboost":       "XGBoost",
    "lightgbm":      "LightGBM",
    "bilstm":        "BiLSTM",
    "gru":           "GRU",
    "transformer":   "Transformer",
    "emd_lstm":      "EMD-LSTM",
    "buy_hold":      "Buy-and-Hold",
    "random":        "Random Signal",
}

MODEL_ORDER = ["qk_svr", "rbf_svm", "laplacian_svm", "xgboost",
               "lightgbm", "bilstm", "gru", "transformer", "emd_lstm",
               "buy_hold", "random"]

FALLBACK_REGIMES = {
    "pre_crisis":   ("2019-01-01", "2021-12-31"),
    "crisis_onset": ("2022-01-01", "2022-06-30"),
    "peak_crisis":  ("2022-07-01", "2023-06-30"),
    "post_crisis":  ("2023-07-01", "2023-12-31"),
}


# ── Helpers ───────────────────────────────────────────────────────
def load_parquet(path: Path) -> Optional[pd.DataFrame]:
    """BUG 6 FIX: fallback CSV."""
    if path.exists():
        return pd.read_parquet(path)
    csv = path.with_suffix(".csv")
    if csv.exists():
        logging.info(f"Parquet not found → using CSV: {csv.name}")
        return pd.read_csv(csv)
    return None


def to_scalar(val) -> float:
    """
    BUG 2+3 FIX: convert scalar, list, or ndarray → single float.
    After 03d_merge_fixed, y_true/y_pred are scalars.
    Handles backward compat if someone passes old list format.
    """
    if isinstance(val, (int, float, np.floating, np.integer)):
        return float(val)
    arr = np.asarray(val, dtype=float).reshape(-1)
    return float(arr[0]) if len(arr) > 0 else float("nan")


def load_break_dates() -> Dict:
    bp = CONFIG_DIR / "break_dates.json"
    if not bp.exists():
        logging.warning("break_dates.json not found → fallback regimes.")
        return FALLBACK_REGIMES
    with open(bp) as f:
        data = json.load(f)
    breaks = sorted(data.get("break_dates", []))
    if len(breaks) < 2:
        return FALLBACK_REGIMES
    return {
        "pre_crisis":   ("2019-01-01", breaks[0]),
        "crisis_onset": (breaks[0], breaks[1]),
        "peak_crisis":  (breaks[1], breaks[2] if len(breaks) > 2 else "2023-06-30"),
        "post_crisis":  (breaks[2] if len(breaks) > 2 else "2023-07-01", "2024-12-31"),
    }


def load_predictions() -> pd.DataFrame:
    """
    BUG 1 FIX: chỉ đọc benchmark_predictions.parquet (QK-SVR đã được merge bởi 03d).
    Standardize y_pred column name.
    """
    df = load_parquet(RESULTS_DIR / "benchmark_predictions.parquet")
    if df is None:
        raise FileNotFoundError(
            "benchmark_predictions.parquet not found.\n"
            "Run 03d_merge_fixed.py first."
        )

    # Standardize y_pred column (PERF FIX: handle both names)
    if "y_pred_mean" in df.columns and "y_pred" not in df.columns:
        df = df.rename(columns={"y_pred_mean": "y_pred"})

    # Ensure datetime
    for col in ["test_start", "test_end"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    if "qk_svr" not in df["model"].unique():
        logging.warning("⚠️ qk_svr not in benchmark_predictions — run 03d_merge_fixed.py")

    logging.info(f"Loaded: {len(df)} rows | Models: {sorted(df['model'].unique())}")
    return df


# ── Trading Metrics ───────────────────────────────────────────────
def compute_sharpe(pnl: np.ndarray) -> float:
    if len(pnl) < 2 or pnl.std() < 1e-10:
        return float("nan")
    return float(pnl.mean() / pnl.std() * np.sqrt(TRADING_DAYS_PER_YEAR))


def compute_max_drawdown(cum: np.ndarray) -> float:
    if len(cum) == 0:
        return float("nan")
    peak = np.maximum.accumulate(cum)
    dd   = (cum - peak) / (np.abs(peak) + 1e-10)
    return float(dd.min() * 100)


def compute_calmar(ann_ret: float, max_dd: float) -> float:
    if abs(max_dd) < 1e-10:
        return float("nan")
    return float(ann_ret / abs(max_dd))


def compute_win_loss(pnl: np.ndarray) -> float:
    wins = pnl[pnl > 0]; losses = pnl[pnl < 0]
    if len(wins) == 0 or len(losses) == 0:
        return float("nan")
    return float(wins.mean() / abs(losses.mean()))


def trading_metrics(y_true: np.ndarray, y_pred: np.ndarray, strategy: str = "model") -> Dict:
    n = len(y_true)
    if strategy == "model":
        position = np.sign(y_pred)
    elif strategy == "buy_hold":
        position = np.ones(n)
    elif strategy == "random":
        position = np.random.default_rng(RANDOM_SEED).choice([-1.0, 1.0], size=n)
    else:
        raise ValueError(f"Unknown strategy: {strategy}")

    pnl         = position * y_true
    cum         = np.cumsum(pnl)
    total_ret   = float(cum[-1]) * 100 if n > 0 else float("nan")
    ann_ret     = total_ret * (TRADING_DAYS_PER_YEAR / max(n, 1))
    sharpe      = compute_sharpe(pnl)
    max_dd      = compute_max_drawdown(cum)
    calmar      = compute_calmar(ann_ret, max_dd)
    win_loss    = compute_win_loss(pnl)

    mask        = y_true != 0
    hit_rate    = (float(np.mean(np.sign(y_true[mask]) == position[mask]) * 100)
                   if mask.sum() > 0 else float("nan"))

    return {
        "total_return_pct":    round(total_ret, 4),
        "ann_return_pct":      round(ann_ret, 4),
        "sharpe_ratio":        round(sharpe, 4) if not np.isnan(sharpe) else float("nan"),
        "max_drawdown_pct":    round(max_dd, 4) if not np.isnan(max_dd) else float("nan"),
        "calmar_ratio":        round(calmar, 4) if not np.isnan(calmar) else float("nan"),
        "hit_rate_pct":        round(hit_rate, 2) if not np.isnan(hit_rate) else float("nan"),
        "win_loss_ratio":      round(win_loss, 4) if not np.isnan(win_loss) else float("nan"),
        "n_trades":            int(np.sum(position != 0)),
        "n_obs":               n,
    }


# ── Main simulation ───────────────────────────────────────────────
def run_simulation(df_preds: pd.DataFrame, regimes: Dict) -> pd.DataFrame:
    """
    BUG 2+5 FIX: use to_scalar() for y_true/y_pred;
    vectorized regime masking with pd.Timestamp comparisons.
    """
    records = []

    for horizon in sorted(df_preds["horizon"].unique()):
        df_h = df_preds[df_preds["horizon"] == horizon]
        model_series: Dict = {}

        for model in df_h["model"].unique():
            df_m = df_h[df_h["model"] == model].sort_values("test_start")
            if len(df_m) == 0:
                continue

            # BUG 2 FIX: collect scalars → arrays
            y_true_arr = np.array([to_scalar(v) for v in df_m["y_true"]])
            y_pred_arr = np.array([to_scalar(v) for v in df_m["y_pred"]])
            dates_arr  = df_m["test_start"].values  # numpy datetime64

            model_series[model] = {
                "y_true": y_true_arr,
                "y_pred": y_pred_arr,
                "dates":  dates_arr,
            }

        if not model_series:
            continue

        ref_key  = list(model_series.keys())[0]
        y_true_r = model_series[ref_key]["y_true"]
        dates_r  = model_series[ref_key]["dates"]

        # Baselines
        model_series["buy_hold"] = {
            "y_true": y_true_r,
            "y_pred": np.ones(len(y_true_r)),
            "dates":  dates_r,
        }
        model_series["random"] = {
            "y_true": y_true_r,
            "y_pred": np.random.default_rng(RANDOM_SEED).choice([-1.0, 1.0], size=len(y_true_r)),
            "dates":  dates_r,
        }

        for model_key, series in model_series.items():
            y_true = series["y_true"]
            y_pred = series["y_pred"]
            dates  = pd.to_datetime(series["dates"])

            strategy = ("buy_hold" if model_key == "buy_hold"
                        else "random" if model_key == "random"
                        else "model")

            # Overall
            m = trading_metrics(y_true, y_pred, strategy)
            m.update({"model": model_key, "horizon": horizon, "regime": "overall"})
            records.append(m)

            # Per regime — BUG 5 FIX: vectorized masking
            for regime_name, (r_start, r_end) in regimes.items():
                mask = (dates >= pd.Timestamp(r_start)) & (dates <= pd.Timestamp(r_end))
                # mask có thể là numpy array hoặc pandas Series tùy source
                if hasattr(mask, 'values'):
                    mask = mask.values
                mask = mask.astype(bool)
                if mask.sum() < 5:
                    continue
                m_r = trading_metrics(y_true[mask], y_pred[mask], strategy)
                m_r.update({"model": model_key, "horizon": horizon, "regime": regime_name})
                records.append(m_r)

        logging.info(f"  H={horizon}: {len(model_series)} models done")

    return pd.DataFrame(records)


def build_cumulative_returns(df_preds: pd.DataFrame) -> pd.DataFrame:
    """BUG 3 FIX: use to_scalar() for y_true/y_pred."""
    records = []
    for horizon in sorted(df_preds["horizon"].unique()):
        df_h = df_preds[df_preds["horizon"] == horizon]
        for model in df_h["model"].unique():
            df_m = df_h[df_h["model"] == model].sort_values("test_start")
            cum  = 0.0
            for _, row in df_m.iterrows():
                y_true = np.array([to_scalar(row["y_true"])])   # BUG 3 FIX
                y_pred = np.array([to_scalar(row["y_pred"])])
                pnl    = np.sign(y_pred) * y_true
                cum   += float(pnl.sum())
                records.append({
                    "model":      model,
                    "horizon":    horizon,
                    "fold_id":    row["fold_id"],
                    "test_start": row.get("test_start"),
                    "fold_pnl":   float(pnl.sum()),
                    "cum_return": cum,
                })
    return pd.DataFrame(records)


# ── XLSX Export ───────────────────────────────────────────────────
def write_table_a1(df_sim: pd.DataFrame, out_path: Path) -> None:
    display = df_sim[
        (df_sim["regime"].isin(["overall", "peak_crisis"])) &
        (df_sim["horizon"].isin([1, 22]))
    ].copy()
    display["model_display"] = display["model"].map(lambda x: MODEL_NAMES.get(x, x))
    display["Horizon"] = display["horizon"].apply(lambda h: f"H={h}")
    display["Regime"]  = display["regime"].str.replace("_", " ").str.title()

    col_map = {
        "model_display":    "Model",
        "Horizon":          "Horizon",
        "Regime":           "Regime",
        "sharpe_ratio":     "Sharpe Ratio",
        "total_return_pct": "Total Return (%)",
        "max_drawdown_pct": "Max Drawdown (%)",
        "calmar_ratio":     "Calmar Ratio",
        "hit_rate_pct":     "Hit Rate (%)",
        "win_loss_ratio":   "Win/Loss",
        "n_trades":         "N Trades",
    }
    display = display.rename(columns=col_map)
    cols    = list(col_map.values())
    display = display[cols]

    order_map = {MODEL_NAMES.get(m, m): i for i, m in enumerate(MODEL_ORDER)}
    display["_o"] = display["Model"].map(lambda x: order_map.get(x, 99))
    display = display.sort_values(["Regime", "Horizon", "_o"]).drop(columns=["_o"])

    for col in ["Sharpe Ratio","Total Return (%)","Max Drawdown (%)","Calmar Ratio","Win/Loss"]:
        display[col] = display[col].apply(lambda v: f"{v:.4f}" if pd.notna(v) else "—")
    display["Hit Rate (%)"] = display["Hit Rate (%)"].apply(
        lambda v: f"{v:.2f}" if pd.notna(v) else "—")

    wb = Workbook()
    ws = wb.active
    ws.title = "Table A1"
    ws.sheet_view.showGridLines = False
    ncols = len(cols)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = "Table A1. Trading Simulation Results"
    ws["A1"].font      = Font(name="Times New Roman", bold=True, size=11)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    for col, name in enumerate(cols, 1):
        c = ws.cell(row=3, column=col, value=name)
        c.font      = Font(name="Times New Roman", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", start_color="1F4E79", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = Border(top=Side(style="medium"), bottom=Side(style="thin"))
    ws.row_dimensions[3].height = 30

    for i, (_, row) in enumerate(display.iterrows()):
        r     = i + 4
        is_qk = row["Model"] == "QK-SVR"
        bg    = "FFF3E0" if is_qk else ("F7F9FC" if i % 2 == 1 else "FFFFFF")
        for col, key in enumerate(cols, 1):
            c = ws.cell(row=r, column=col, value=row[key])
            c.font      = Font(name="Times New Roman", bold=is_qk, size=10)
            c.fill      = PatternFill("solid", start_color=bg, fgColor=bg)
            c.alignment = (Alignment(horizontal="left", vertical="center")
                           if col == 1
                           else Alignment(horizontal="center", vertical="center"))

    for col in range(1, ncols+1):
        ws.cell(row=ws.max_row, column=col).border = Border(bottom=Side(style="medium"))

    note_row = ws.max_row + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=ncols)
    # BUG 4 FIX: removed duplicate "Note. " prefix
    note_text = (
        "Long-short strategy: position = sign(forecast). "
        "Zero transaction cost. Unit position size. "
        "Sharpe ratio annualised using √252 scaling factor. "
        "Max Drawdown = maximum peak-to-trough decline in cumulative log-return. "
        "Calmar = annualised return / |max drawdown|. "
        "Hit Rate = directional accuracy (%). "
        "Buy-and-Hold: position = +1 always. Random Signal: position = random ±1 (seed=42). "
        "QK-SVR row highlighted."
    )
    c = ws.cell(row=note_row, column=1, value=f"Note. {note_text}")
    c.font      = Font(name="Times New Roman", italic=True, size=9)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[note_row].height = 60

    for col in ws.columns:
        w = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w+2, 10), 30)

    wb.save(out_path)
    logging.info(f"  Saved: {out_path.name}")


# ── Main ──────────────────────────────────────────────────────────
def main():
    df_preds = load_predictions()
    regimes  = load_break_dates()

    logging.info("\n[1/3] Running trading simulation...")
    df_sim = run_simulation(df_preds, regimes)
    df_sim.to_csv(RESULTS_DIR / "trading_simulation.csv", index=False)
    logging.info(f"  Saved trading_simulation.csv ({len(df_sim)} rows)")

    logging.info("\n[2/3] Building cumulative returns...")
    df_cum = build_cumulative_returns(df_preds)
    df_cum.to_csv(RESULTS_DIR / "trading_cumulative_returns.csv", index=False)
    logging.info(f"  Saved trading_cumulative_returns.csv ({len(df_cum)} rows)")

    logging.info("\n[3/3] Exporting Table A1...")
    write_table_a1(df_sim, TABLE_DIR / "Table_A1_Trading_Simulation.xlsx")

    # Console summary
    print(f"\n{'='*65}")
    print("  TRADING SIMULATION SUMMARY")
    print(f"{'='*65}")
    order_map = {MODEL_NAMES.get(m, m): i for i, m in enumerate(MODEL_ORDER)}

    for horizon in [1, 5, 22]:
        sub = df_sim[(df_sim["regime"]=="overall") & (df_sim["horizon"]==horizon)].copy()
        if len(sub) == 0:
            continue
        sub["Model"] = sub["model"].map(lambda x: MODEL_NAMES.get(x, x))
        sub["_o"]    = sub["Model"].map(lambda x: order_map.get(x, 99))
        sub = sub.sort_values("_o")
        print(f"\n  Overall H={horizon}:")
        print(sub[["Model","sharpe_ratio","total_return_pct","hit_rate_pct"]]
              .to_string(index=False))

    print(f"\n{'='*65}")
    print(f"  Outputs: {RESULTS_DIR}")
    print(f"  Table:   {TABLE_DIR / 'Table_A1_Trading_Simulation.xlsx'}")
    print(f"{'='*65}")


if __name__ == "__main__":
    main()
